"""
Diagnosis Tab - Main Controller
Date & Time-Based Report Sheet with charts and export functionality
"""

import tkinter as tk

from frontend.utils.config import AppConfig
from ..utils.styles import Colors, Fonts
from ..utils.db_error_handler import DatabaseErrorHandler
from ..utils.debug_logger import log_error, log_info
from .report_data_table import ReportDataTable
from .control_panel import ControlPanel
from .action_panel import ActionPanel
from .status_chart import StatusChart
from .defectwise_chart import DefectwiseChart


class DiagnosisTab:
    """Diagnosis tab for generating and viewing reports with charts."""
    
    def __init__(self, parent, app_instance):
        """
        Initialize the diagnosis tab.
        
        Args:
            parent: Parent frame (tab)
            app_instance: Reference to main WelVisionApp instance
        """
        self.parent = parent
        self.app = app_instance
        
        # Components
        self.report_data_table = None
        self.control_panel = None
        self.action_panel = None
        self.status_chart = None
        self.defectwise_chart = None
        
        # Current report data
        self.current_data = []
        
    def setup(self):
        """Setup the diagnosis tab UI."""
        try:
            log_info("diagnosis", "Setting up Diagnosis tab")
            
            # Main container with dark blue background
            main_container = tk.Frame(self.parent, bg=Colors.PRIMARY_BG)
            main_container.pack(fill=tk.BOTH, expand=True)
            
            # Header frame for title
            header_frame = tk.Frame(main_container, bg=Colors.PRIMARY_BG)
            header_frame.pack(fill=tk.X, pady=(10, 10))
            
            # Title (centered)
            title_label = tk.Label(
                header_frame,
                text="Date & Time-Based Report Sheet",
                font=Fonts.TITLE,
                fg=Colors.WHITE,
                bg=Colors.PRIMARY_BG
            )
            title_label.pack()
            
            # Top section: Report Data Table + Controls + Actions (increased height for visibility)
            top_frame = tk.Frame(main_container, bg=Colors.PRIMARY_BG, height=280)
            top_frame.pack(fill=tk.X, padx=10, pady=5)
            top_frame.pack_propagate(False)  # Prevent frame from shrinking
            
            # Report Data Table (left side)
            self.report_data_table = ReportDataTable(top_frame, self)
            self.report_data_table.create()
            
            # Right panel: Controls + Actions (fixed width for better control display)
            right_panel = tk.Frame(top_frame, bg=Colors.PRIMARY_BG, width=320)
            right_panel.pack(side=tk.LEFT, fill=tk.Y, padx=(10, 0))
            right_panel.pack_propagate(False)  # Maintain fixed width
            
            # Control Panel
            self.control_panel = ControlPanel(right_panel, self)
            self.control_panel.create()
            
            # Action Panel
            self.action_panel = ActionPanel(right_panel, self)
            self.action_panel.create()
            
            # Bottom section: Charts (side by side with more height)
            charts_frame = tk.Frame(main_container, bg=Colors.PRIMARY_BG)
            charts_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
            
            # LAZY LOADING: Create charts only when data is loaded
            # Show placeholder message initially
            self.charts_placeholder = tk.Label(
                charts_frame,
                text="📊 Generate a report to view charts",
                font=Fonts.SUBTITLE,
                fg="#888888",
                bg=Colors.PRIMARY_BG,
                pady=50
            )
            self.charts_placeholder.pack(fill=tk.BOTH, expand=True)
            
            log_info("diagnosis", "Diagnosis tab setup completed")
            
        except Exception as e:
            log_error("diagnosis", "Failed to setup Diagnosis tab", e)
            raise
        
        # Store charts frame reference for later
        self.charts_frame = charts_frame
    
    def generate_report(self):
        """Generate report based on selected filters."""
        if self.control_panel:
            filters = self.control_panel.get_filters()
            if filters:  # Only proceed if validation passed
                self._fetch_and_display_data(filters)
    
    def save_chart(self):
        """Save the current charts as images."""
        import os
        import subprocess
        from tkinter import messagebox
        
        try:
            # Check if data exists
            if not self.current_data:
                messagebox.showwarning(
                    "No Data",
                    "No data found for the applied filters.\n\n"
                    "Please generate a report first before saving charts."
                )
                return
            
            # Get filters
            filters = self.control_panel.get_filters()
            if not filters:  # Validation failed
                return
            component_type = filters['component_type']
            report_type = filters['report_type']
            from_date = filters['from_date']
            to_date = filters['to_date']
            
            username = os.getlogin()
            
            # Save Status Chart
            if self.status_chart and self.status_chart.figure:
                # Create directory path for Status Chart
                status_path = f"C:\\Users\\{username}\\Desktop\\Diagnosis\\Charts\\{component_type}\\{report_type}\\Status"
                os.makedirs(status_path, exist_ok=True)
                
                # Create filename
                status_filename = f"{component_type}_{report_type}_Status_{from_date}_{to_date}.png"
                status_filepath = os.path.join(status_path, status_filename)
                
                # Save the figure
                self.status_chart.figure.savefig(status_filepath, dpi=300, bbox_inches='tight')
            
            # Save Defectwise Chart
            if self.defectwise_chart and self.defectwise_chart.figure:
                # Create directory path for Defectwise Chart
                defectwise_path = f"C:\\Users\\{username}\\Desktop\\Diagnosis\\Charts\\{component_type}\\{report_type}\\Defectwise"
                os.makedirs(defectwise_path, exist_ok=True)
                
                # Create filename
                defectwise_filename = f"{component_type}_{report_type}_Defectwise_{from_date}_{to_date}.png"
                defectwise_filepath = os.path.join(defectwise_path, defectwise_filename)
                
                # Save the figure
                self.defectwise_chart.figure.savefig(defectwise_filepath, dpi=300, bbox_inches='tight')
            
            # Create custom dialog with "Open Location" option
            result = messagebox.askquestion(
                "Charts Saved", 
                f"Charts saved successfully!\n\n"
                f"Status Chart:\n{status_filepath}\n\n"
                f"Defectwise Chart:\n{defectwise_filepath}\n\n"
                f"Do you want to open the folder location?",
                icon='info'
            )
            
            if result == 'yes':
                # Open the parent directory (Charts folder)
                charts_folder = os.path.dirname(status_filepath)
                subprocess.run(['explorer', charts_folder])
            
        except Exception as e:
            print(f"❌ Error saving charts: {e}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Save Error", f"Failed to save charts:\n{str(e)}")
    
    def export_to_excel(self):
        """Export report data to Excel file."""
        import os
        import subprocess
        import pandas as pd
        from tkinter import messagebox
        from openpyxl.styles import Font, PatternFill
        
        if not self.current_data:
            messagebox.showwarning(
                "No Data",
                "No data found for the applied filters.\n\n"
                "Please generate a report first before exporting."
            )
            return
        
        try:
            
            # Get filters
            filters = self.control_panel.get_filters()
            if not filters:  # Validation failed
                return
            
            component_type = filters['component_type']
            report_type = filters['report_type']
            from_date = filters['from_date']
            to_date = filters['to_date']
            
            # Create directory path
            username = os.getlogin()
            base_path = f"C:\\Users\\{username}\\Desktop\\Diagnosis\\Report\\{component_type}\\{report_type}"
            os.makedirs(base_path, exist_ok=True)
            
            # Create filename
            filename = f"{component_type}_{report_type}_{from_date}_{to_date}.xlsx"
            filepath = os.path.join(base_path, filename)
            
            # Convert data to DataFrame and add S.No column
            df = pd.DataFrame(self.current_data)
            
            # Insert S.No column at the beginning
            df.insert(0, 'S.No', range(1, len(df) + 1))
            
            # Calculate sum row for numeric columns
            sum_row = {}
            total_inspected = 0
            total_accepted = 0
            
            for col in df.columns:
                if col == 'S.No':
                    sum_row[col] = 'Total'
                elif col in ['Component Type', 'Employee ID', 'Report Date', 'Start Time', 'End Time']:
                    sum_row[col] = ''
                elif col == 'Acceptance Rate':
                    # Calculate percentage for sum row
                    if total_inspected > 0:
                        acceptance_rate = (total_accepted / total_inspected) * 100
                        sum_row[col] = f"{acceptance_rate:.2f}%"
                    else:
                        sum_row[col] = "0.00%"
                else:
                    # Sum numeric columns
                    try:
                        sum_value = df[col].sum()
                        sum_row[col] = sum_value
                        
                        # Track inspected and accepted for percentage calculation
                        if 'Inspected' in col:
                            total_inspected = sum_value
                        elif 'Accepted' in col:
                            total_accepted = sum_value
                    except:
                        sum_row[col] = ''
            
            # Add sum row to DataFrame
            df = pd.concat([df, pd.DataFrame([sum_row])], ignore_index=True)
            
            # Export to Excel
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Report')
                
                # Get the worksheet
                worksheet = writer.sheets['Report']
                
                # Auto-adjust column widths
                for idx, col in enumerate(df.columns):
                    max_length = max(
                        df[col].astype(str).apply(len).max(),
                        len(str(col))
                    )
                    worksheet.column_dimensions[chr(65 + idx)].width = min(max_length + 2, 50)
                
                # Style the sum row (last row) - Blue color and bold font
                last_row_idx = len(df) + 1  # +1 because Excel rows are 1-indexed and we have header
                blue_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
                white_font = Font(bold=True, color="FFFFFF")
                
                for col_idx in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=last_row_idx, column=col_idx)
                    cell.fill = blue_fill
                    cell.font = white_font
            
            # Create custom dialog with "Open Location" option
            result = messagebox.askquestion(
                "Export Success",
                f"Report exported successfully!\n\n"
                f"File Location:\n{filepath}\n\n"
                f"Do you want to open the folder location?",
                icon='info'
            )
            
            if result == 'yes':
                # Open the folder containing the exported file
                subprocess.run(['explorer', base_path])
            
        except Exception as e:
            print(f"❌ Error exporting to Excel: {e}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Export Error", f"Failed to export report:\n{str(e)}")
    
    def clear_database(self):
        """Clear database entries for the applied filters (Admin and Super Admin only)."""
        from tkinter import messagebox
        
        # Double-check user role for security
        if not hasattr(self.app, 'current_role') or self.app.current_role not in ['Admin', 'Super Admin']:
            messagebox.showerror(
                "Access Denied",
                "⛔ You do not have permission to clear database entries.\n\n"
                "This action is restricted to Admin and Super Admin users only."
            )
            return
        
        # Get filters
        filters = self.control_panel.get_filters()
        if not filters:  # Validation failed
            return
        
        component_type = filters['component_type']
        report_type = filters['report_type']
        from_date = filters['from_date']
        to_date = filters['to_date']
        
        # Confirmation dialog with detailed information
        confirm_message = (
            f"⚠️ WARNING: Database Deletion\n\n"
            f"You are about to permanently delete data from the database:\n\n"
            f"Report Type: {report_type}\n"
            f"Component Type: {component_type}\n"
            f"Date Range: {from_date} to {to_date}\n\n"
            f"This action CANNOT be undone!\n\n"
            f"Are you sure you want to continue?"
        )
        
        response = messagebox.askyesno(
            "Confirm Database Deletion",
            confirm_message,
            icon='warning'
        )
        
        if not response:
            return  # User cancelled
        
        # Second confirmation for extra safety
        final_confirm = messagebox.askyesno(
            "Final Confirmation",
            "⚠️ LAST WARNING ⚠️\n\n"
            "This will permanently delete the selected data.\n\n"
            "Click YES to proceed with deletion.\n"
            "Click NO to cancel.",
            icon='warning'
        )
        
        if not final_confirm:
            return  # User cancelled
        
        # Proceed with deletion
        try:
            import mysql.connector
            
            connection = mysql.connector.connect(
                host=AppConfig.DB_HOST,
                user=AppConfig.DB_USER,
                password=AppConfig.DB_PASSWORD,
                database=AppConfig.DB_DATABASE
            )
            
            cursor = connection.cursor()
            
            deleted_count = 0
            
            # Determine if we need to filter by component type or delete all
            use_component_filter = (component_type != "All")
            
            if report_type == 'BF':
                # Delete from BF table only
                if use_component_filter:
                    delete_query = """
                        DELETE FROM bf_roller_tracking
                        WHERE roller_type = %s AND report_date BETWEEN %s AND %s
                    """
                    cursor.execute(delete_query, (component_type, from_date, to_date))
                else:
                    # Delete all roller types
                    delete_query = """
                        DELETE FROM bf_roller_tracking
                        WHERE report_date BETWEEN %s AND %s
                    """
                    cursor.execute(delete_query, (from_date, to_date))
                deleted_count = cursor.rowcount
                
            elif report_type == 'OD':
                # Delete from OD table only
                if use_component_filter:
                    delete_query = """
                        DELETE FROM od_roller_tracking
                        WHERE roller_type = %s AND report_date BETWEEN %s AND %s
                    """
                    cursor.execute(delete_query, (component_type, from_date, to_date))
                else:
                    # Delete all roller types
                    delete_query = """
                        DELETE FROM od_roller_tracking
                        WHERE report_date BETWEEN %s AND %s
                    """
                    cursor.execute(delete_query, (from_date, to_date))
                deleted_count = cursor.rowcount
                
            else:  # Overall - delete from both tables
                # Delete from BF table
                if use_component_filter:
                    bf_delete_query = """
                        DELETE FROM bf_roller_tracking
                        WHERE roller_type = %s AND report_date BETWEEN %s AND %s
                    """
                    cursor.execute(bf_delete_query, (component_type, from_date, to_date))
                else:
                    # Delete all roller types
                    bf_delete_query = """
                        DELETE FROM bf_roller_tracking
                        WHERE report_date BETWEEN %s AND %s
                    """
                    cursor.execute(bf_delete_query, (from_date, to_date))
                bf_count = cursor.rowcount
                
                # Delete from OD table
                if use_component_filter:
                    od_delete_query = """
                        DELETE FROM od_roller_tracking
                        WHERE roller_type = %s AND report_date BETWEEN %s AND %s
                    """
                    cursor.execute(od_delete_query, (component_type, from_date, to_date))
                else:
                    # Delete all roller types
                    od_delete_query = """
                        DELETE FROM od_roller_tracking
                        WHERE report_date BETWEEN %s AND %s
                    """
                    cursor.execute(od_delete_query, (from_date, to_date))
                od_count = cursor.rowcount
                
                deleted_count = bf_count + od_count
            
            # Commit the transaction
            connection.commit()
            
            cursor.close()
            connection.close()
            
            # Show success message
            messagebox.showinfo(
                "Database Cleared",
                f"✅ Successfully deleted {deleted_count} record(s) from database.\n\n"
                f"Report Type: {report_type}\n"
                f"Component Type: {component_type}\n"
                f"Date Range: {from_date} to {to_date}"
            )
            
            # Clear current data and refresh the view
            self.current_data = []
            if self.report_data_table:
                self.report_data_table.update_data([], report_type)
            
            # Clear charts
            self._update_charts([], report_type)
            
        except Exception as e:
            print(f"❌ Error clearing database: {e}")
            import traceback
            traceback.print_exc()
            DatabaseErrorHandler.handle_db_error(e, self.parent, "clearing database entries")
    
    
    def _fetch_and_display_data(self, filters):
        """
        Fetch data from database and display in table and charts.
        
        Args:
            filters: Dictionary with report_type, component_type, from_date, to_date
        """
        try:
            import mysql.connector
            from datetime import datetime
            
            connection = mysql.connector.connect(
                host=AppConfig.DB_HOST,
                user=AppConfig.DB_USER,
                password=AppConfig.DB_PASSWORD,
                database=AppConfig.DB_DATABASE
            )
            
            cursor = connection.cursor()
            
            report_type = filters['report_type']
            component_type = filters['component_type']  
            from_date = filters['from_date']
            to_date = filters['to_date']
            
            all_data = []
            
            # Determine if we need to filter by component type or get all
            use_component_filter = (component_type != "All")
            
            if report_type == 'BF':
                # Query BF table only
                if use_component_filter:
                    query = """
                        SELECT 
                            roller_type, employee_id, report_date, start_time, end_time,
                            total_inspected, total_accepted, total_rejected,
                            total_rust, total_dent, total_damage, 
                            total_high_head, total_low_head, others
                        FROM bf_roller_tracking
                        WHERE roller_type = %s AND report_date BETWEEN %s AND %s
                        ORDER BY report_date DESC, start_time DESC
                    """
                    cursor.execute(query, (component_type, from_date, to_date))
                else:
                    # Get all roller types
                    query = """
                        SELECT 
                            roller_type, employee_id, report_date, start_time, end_time,
                            total_inspected, total_accepted, total_rejected,
                            total_rust, total_dent, total_damage, 
                            total_high_head, total_low_head, others
                        FROM bf_roller_tracking
                        WHERE report_date BETWEEN %s AND %s
                        ORDER BY report_date DESC, start_time DESC
                    """
                    cursor.execute(query, (from_date, to_date))
                
                for row in cursor.fetchall():
                    comp_type, emp_id, rep_date, start, end, inspected, accepted, rejected, \
                    rust, dent, damage, high_head, low_head, others = row
                    
                    acc_rate = (accepted / inspected * 100) if inspected > 0 else 0
                    
                    all_data.append({
                        'Component Type': comp_type,
                        'Employee ID': emp_id,
                        'Report Date': rep_date.strftime('%Y-%m-%d'),
                        'Start Time': str(start),
                        'End Time': str(end),
                        'BF Inspected': inspected,
                        'BF Accepted': accepted,
                        'BF Rejected': rejected,
                        'Acceptance Rate': f"{acc_rate:.2f}%",
                        'Rust': rust,
                        'Damage': damage,
                        'Dent': dent,
                        'High Head': high_head,
                        'Down Head': low_head,
                        'Others': others
                    })
            
            elif report_type == 'OD':
                # Query OD table only
                if use_component_filter:
                    query = """
                        SELECT 
                            roller_type, employee_id, report_date, start_time, end_time,
                            total_inspected, total_accepted, total_rejected,
                            total_rust, total_dent, total_damage,
                            total_damage_on_end, total_spherical, others
                        FROM od_roller_tracking
                        WHERE roller_type = %s AND report_date BETWEEN %s AND %s
                        ORDER BY report_date DESC, start_time DESC
                    """
                    cursor.execute(query, (component_type, from_date, to_date))
                else:
                    # Get all roller types
                    query = """
                        SELECT 
                            roller_type, employee_id, report_date, start_time, end_time,
                            total_inspected, total_accepted, total_rejected,
                            total_rust, total_dent, total_damage,
                            total_damage_on_end, total_spherical, others
                        FROM od_roller_tracking
                        WHERE report_date BETWEEN %s AND %s
                        ORDER BY report_date DESC, start_time DESC
                    """
                    cursor.execute(query, (from_date, to_date))
                
                for row in cursor.fetchall():
                    comp_type, emp_id, rep_date, start, end, inspected, accepted, rejected, \
                    rust, dent, damage, damage_on_end, spherical, others = row
                    
                    acc_rate = (accepted / inspected * 100) if inspected > 0 else 0
                    
                    all_data.append({
                        'Component Type': comp_type,
                        'Employee ID': emp_id,
                        'Report Date': rep_date.strftime('%Y-%m-%d'),
                        'Start Time': str(start),
                        'End Time': str(end),
                        'OD Inspected': inspected,
                        'OD Accepted': accepted,
                        'OD Rejected': rejected,
                        'Acceptance Rate': f"{acc_rate:.2f}%",
                        'Rust': rust,
                        'Damage': damage,
                        'Dent': dent,
                        'Damage on End': damage_on_end,
                        'Spherical Mark': spherical,
                        'Others': others
                    })
            
            else:  # Overall
                # Query both tables and combine
                if use_component_filter:
                    bf_query = """
                        SELECT 
                            roller_type, employee_id, report_date, start_time, end_time,
                            total_inspected, total_accepted, total_rejected
                        FROM bf_roller_tracking
                        WHERE roller_type = %s AND report_date BETWEEN %s AND %s
                        ORDER BY report_date DESC, start_time DESC
                    """
                    cursor.execute(bf_query, (component_type, from_date, to_date))
                else:
                    # Get all roller types
                    bf_query = """
                        SELECT 
                            roller_type, employee_id, report_date, start_time, end_time,
                            total_inspected, total_accepted, total_rejected
                        FROM bf_roller_tracking
                        WHERE report_date BETWEEN %s AND %s
                        ORDER BY report_date DESC, start_time DESC
                    """
                    cursor.execute(bf_query, (from_date, to_date))
                bf_rows = cursor.fetchall()
                
                if use_component_filter:
                    od_query = """
                        SELECT 
                            roller_type, employee_id, report_date, start_time, end_time,
                            total_inspected, total_accepted, total_rejected
                        FROM od_roller_tracking
                        WHERE roller_type = %s AND report_date BETWEEN %s AND %s
                        ORDER BY report_date DESC, start_time DESC
                    """
                    cursor.execute(od_query, (component_type, from_date, to_date))
                else:
                    # Get all roller types
                    od_query = """
                        SELECT 
                            roller_type, employee_id, report_date, start_time, end_time,
                            total_inspected, total_accepted, total_rejected
                        FROM od_roller_tracking
                        WHERE report_date BETWEEN %s AND %s
                        ORDER BY report_date DESC, start_time DESC
                    """
                    cursor.execute(od_query, (from_date, to_date))
                od_rows = cursor.fetchall()
                
                # Create a dictionary to group by date and employee
                combined_data = {}
                
                for row in bf_rows:
                    comp_type, emp_id, rep_date, start, end, inspected, accepted, rejected = row
                    key = (emp_id, rep_date, start)
                    
                    if key not in combined_data:
                        combined_data[key] = {
                            'Component Type': comp_type,
                            'Employee ID': emp_id,
                            'Report Date': rep_date.strftime('%Y-%m-%d'),
                            'Start Time': str(start),
                            'End Time': str(end),
                            'BF Inspected': inspected,
                            'BF Accepted': accepted,
                            'BF Rejected': rejected,
                            'OD Inspected': 0,
                            'OD Accepted': 0,
                            'OD Rejected': 0
                        }
                    else:
                        combined_data[key]['BF Inspected'] = inspected
                        combined_data[key]['BF Accepted'] = accepted
                        combined_data[key]['BF Rejected'] = rejected
                
                for row in od_rows:
                    comp_type, emp_id, rep_date, start, end, inspected, accepted, rejected = row
                    key = (emp_id, rep_date, start)
                    
                    if key not in combined_data:
                        combined_data[key] = {
                            'Component Type': comp_type,
                            'Employee ID': emp_id,
                            'Report Date': rep_date.strftime('%Y-%m-%d'),
                            'Start Time': str(start),
                            'End Time': str(end),
                            'BF Inspected': 0,
                            'BF Accepted': 0,
                            'BF Rejected': 0,
                            'OD Inspected': inspected,
                            'OD Accepted': accepted,
                            'OD Rejected': rejected
                        }
                    else:
                        combined_data[key]['OD Inspected'] = inspected
                        combined_data[key]['OD Accepted'] = accepted
                        combined_data[key]['OD Rejected'] = rejected
                
                # Calculate overall values
                for key, data in combined_data.items():
                    data['Overall Inspected'] = data['BF Inspected']
                    data['Overall Accepted'] = data['OD Accepted']
                    data['Overall Rejected'] = data['BF Rejected'] + data['OD Rejected']
                    
                    acc_rate = (data['Overall Accepted'] / data['Overall Inspected'] * 100) if data['Overall Inspected'] > 0 else 0
                    data['Acceptance Rate'] = f"{acc_rate:.2f}%"
                    
                    all_data.append(data)
            
            cursor.close()
            connection.close()
            
            # Update table
            if self.report_data_table:
                self.report_data_table.update_data(all_data, report_type)
            
            # Update charts
            self._update_charts(all_data, report_type)
            
            self.current_data = all_data
            
        except Exception as e:
            print(f"❌ Error fetching report data: {e}")
            import traceback
            traceback.print_exc()
            DatabaseErrorHandler.handle_db_error(e, self.parent, "fetching report data")
    
    def _update_charts(self, data, report_type):
        """Update both charts with the report data based on report type."""
        # LAZY LOADING: Create charts only when needed
        if self.status_chart is None or self.defectwise_chart is None:
            # Remove placeholder
            if hasattr(self, 'charts_placeholder') and self.charts_placeholder:
                self.charts_placeholder.destroy()
            
            # Create charts for the first time
            self.status_chart = StatusChart(self.charts_frame, self)
            self.status_chart.create()
            
            self.defectwise_chart = DefectwiseChart(self.charts_frame, self)
            self.defectwise_chart.create()
        
        # Always update charts, even if data is empty
        if report_type == 'BF':
            # Calculate BF statistics for status chart
            total_inspected = sum(row.get('BF Inspected', 0) for row in data) if data else 0
            total_accepted = sum(row.get('BF Accepted', 0) for row in data) if data else 0
            total_rejected = sum(row.get('BF Rejected', 0) for row in data) if data else 0
            
            status_data = {
                'BF Inspected': total_inspected,
                'BF Accepted': total_accepted,
                'BF Rejected': total_rejected
            }
            
            # Calculate defect statistics
            defect_data = {
                'Rust': sum(row.get('Rust', 0) for row in data) if data else 0,
                'Damage': sum(row.get('Damage', 0) for row in data) if data else 0,
                'Dent': sum(row.get('Dent', 0) for row in data) if data else 0,
                'High Head': sum(row.get('High Head', 0) for row in data) if data else 0,
                'Down Head': sum(row.get('Down Head', 0) for row in data) if data else 0,
                'Others': sum(row.get('Others', 0) for row in data) if data else 0
            }
            
        elif report_type == 'OD':
            # Calculate OD statistics for status chart
            total_inspected = sum(row.get('OD Inspected', 0) for row in data) if data else 0
            total_accepted = sum(row.get('OD Accepted', 0) for row in data) if data else 0
            total_rejected = sum(row.get('OD Rejected', 0) for row in data) if data else 0
            
            status_data = {
                'OD Inspected': total_inspected,
                'OD Accepted': total_accepted,
                'OD Rejected': total_rejected
            }
            
            # Calculate defect statistics
            defect_data = {
                'Rust': sum(row.get('Rust', 0) for row in data) if data else 0,
                'Damage': sum(row.get('Damage', 0) for row in data) if data else 0,
                'Dent': sum(row.get('Dent', 0) for row in data) if data else 0,
                'Damage on End': sum(row.get('Damage on End', 0) for row in data) if data else 0,
                'Spherical Mark': sum(row.get('Spherical Mark', 0) for row in data) if data else 0,
                'Others': sum(row.get('Others', 0) for row in data) if data else 0
            }
            
        else:  # Overall
            # Calculate overall statistics for status chart
            total_inspected = sum(row.get('Overall Inspected', 0) for row in data) if data else 0
            total_accepted = sum(row.get('Overall Accepted', 0) for row in data) if data else 0
            total_rejected = sum(row.get('Overall Rejected', 0) for row in data) if data else 0
            
            status_data = {
                'Overall Inspected': total_inspected,
                'Overall Accepted': total_accepted,
                'Overall Rejected': total_rejected
            }
            
            # Calculate component-wise statistics for defectwise chart
            defect_data = {
                'BF Inspected': sum(row.get('BF Inspected', 0) for row in data) if data else 0,
                'BF Accepted': sum(row.get('BF Accepted', 0) for row in data) if data else 0,
                'BF Rejected': sum(row.get('BF Rejected', 0) for row in data) if data else 0,
                'OD Inspected': sum(row.get('OD Inspected', 0) for row in data) if data else 0,
                'OD Accepted': sum(row.get('OD Accepted', 0) for row in data) if data else 0,
                'OD Rejected': sum(row.get('OD Rejected', 0) for row in data) if data else 0
            }
        
        # Update charts
        if self.status_chart:
            self.status_chart.update_chart(status_data, report_type)
        
        if self.defectwise_chart:
            self.defectwise_chart.update_chart(defect_data, report_type)
